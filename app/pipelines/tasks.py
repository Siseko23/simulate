"""
FreightFlow Nexus - Background Task Pipelines

Pipeline 1 : Supplier Score Recalculation
Pipeline 2 : Shipper Health Score Refresh
Pipeline 3 : Booking Risk Flagging
Pipeline 4 : Quote Expiry ETL
Pipeline 5 : Invoice Auto-Generation
Pipeline 6 : Daily Platform Summary
Pipeline 7 : CSV/Excel Export Generation
"""
import logging
from datetime import datetime, timedelta, date

log = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────────────────────
# These tasks are registered on the Celery instance created in run.py.
# We import `celery` from there at call-time to avoid circular imports.
# ─────────────────────────────────────────────────────────────────────────────

def get_celery():
    from run import celery
    return celery


# ── Pipeline 1: Supplier Score Recalculation ─────────────────────────────────

def recalc_all_supplier_scores():
    """
    Full ETL pipeline:
    Extract  - pull all Active suppliers + their last 60 delivered bookings
    Transform - compute weighted score from on_time_rate, cancellation_rate, rating
    Load      - write new score + score_history record
    """
    from app.models import db, SupplierProfile, Booking, SupplierScoreHistory

    log.info("[PIPELINE] Supplier score recalc started")
    suppliers = SupplierProfile.query.filter_by(status="Active").all()
    updated = 0

    for supplier in suppliers:
        recent_bookings = supplier.bookings\
            .filter(Booking.status.in_(["Delivered", "Cancelled"]))\
            .order_by(Booking.created_at.desc()).limit(60).all()

        if not recent_bookings:
            continue

        total     = len(recent_bookings)
        on_time   = sum(1 for b in recent_bookings if b.status == "Delivered" and b.pod_signed)
        cancelled = sum(1 for b in recent_bookings if b.status == "Cancelled")

        on_time_rate   = (on_time / total) * 100 if total else 100
        cancel_rate    = (cancelled / total) * 100 if total else 0

        # Weighted score: on-time 60%, cancellation penalty 25%, base 15%
        raw_score = (
            (on_time_rate / 100) * 3.0 +
            ((100 - cancel_rate) / 100) * 1.25 +
            0.75  # base
        )
        new_score = max(1.0, min(5.0, round(raw_score, 2)))

        # Only save history if score changed meaningfully
        if abs(new_score - supplier.score) >= 0.01:
            supplier.score       = new_score
            supplier.total_jobs  = total
            supplier.on_time_jobs = on_time
            supplier.cancelled_jobs = cancelled

            hist = SupplierScoreHistory(
                supplier_id=supplier.id,
                score=new_score,
                on_time_rate=round(on_time_rate, 1),
                cancel_rate=round(cancel_rate, 1)
            )
            db.session.add(hist)

            # Auto-suspend if score drops below 3.0
            if new_score < 3.0 and supplier.status == "Active":
                supplier.status = "Suspended"
                from app.services.notifications import push_notification
                push_notification(
                    supplier.user_id,
                    "Account suspended - score below threshold",
                    f"Your score dropped to {new_score}/5.0. Contact support@movement.com.",
                    type="error"
                )
                log.warning(f"[PIPELINE] Supplier {supplier.company_name} auto-suspended (score {new_score})")

            updated += 1

    db.session.commit()
    log.info(f"[PIPELINE] Supplier score recalc complete - {updated}/{len(suppliers)} updated")
    return {"updated": updated, "total": len(suppliers)}


# ── Pipeline 2: Shipper Health Score Refresh ─────────────────────────────────

def refresh_all_health_scores():
    """
    Extract  - all shipper profiles with at least 1 booking
    Transform - run compute_health_score()
    Load      - persist health_score field on ShipperProfile
    """
    from app.models import db, ShipperProfile
    from app.services.ai_engine import compute_health_score

    log.info("[PIPELINE] Health score refresh started")
    shippers = ShipperProfile.query.all()
    updated  = 0

    for shipper in shippers:
        if shipper.bookings.count() == 0:
            continue
        scores = compute_health_score(shipper)
        shipper.health_score = scores["health_score"]
        updated += 1

    db.session.commit()
    log.info(f"[PIPELINE] Health scores refreshed - {updated} shippers")
    return {"updated": updated}


# ── Pipeline 3: Booking Risk Flagging ────────────────────────────────────────

def flag_risky_bookings():
    """
    Extract  - all In Transit bookings
    Transform - compute risk based on supplier score + time overdue
    Load      - set booking.risk_level, push notification if High
    """
    from app.models import db, Booking
    from app.services.notifications import push_notification

    log.info("[PIPELINE] Risk flagging started")
    in_transit = Booking.query.filter(
        Booking.status.in_(["Confirmed","Driver Assigned","Collected","In Transit"])
    ).all()

    flagged = 0
    for booking in in_transit:
        # Risk factors
        risk_score = 0

        # Supplier score
        if booking.supplier:
            if booking.supplier.score < 3.5:   risk_score += 3
            elif booking.supplier.score < 4.0: risk_score += 1

        # Overdue check
        if booking.collection_date:
            days_since = (date.today() - booking.collection_date).days
            if days_since > 3 and booking.status in ("Confirmed","Driver Assigned"):
                risk_score += 2  # not collected yet, overdue
            elif days_since > 5 and booking.status == "Collected":
                risk_score += 1  # in transit longer than expected

        # DC delivery is time-critical
        if booking.destination_type == "DC":
            risk_score += 1

        old_risk = booking.risk_level
        if risk_score >= 4:   booking.risk_level = "High"
        elif risk_score >= 2: booking.risk_level = "Medium"
        else:                  booking.risk_level = "Low"

        if booking.risk_level == "High" and old_risk != "High":
            push_notification(
                booking.shipper.user_id,
                f"Warning High delivery risk - {booking.ref}",
                f"Booking {booking.ref} ({booking.route}) has been flagged as high risk.",
                type="warning", ref_type="booking", ref_id=booking.ref
            )
            flagged += 1

    db.session.commit()
    log.info(f"[PIPELINE] Risk flagging complete - {flagged} bookings flagged High")
    return {"flagged": flagged, "checked": len(in_transit)}


# ── Pipeline 4: Quote Expiry ETL ─────────────────────────────────────────────

def expire_old_quotes():
    """
    Extract  - all Pending quotes older than 48 hours
    Transform - mark as Expired
    Load      - persist, notify supplier
    """
    from app.models import db, Quote
    from app.services.notifications import push_notification

    cutoff  = datetime.utcnow() - timedelta(hours=48)
    expired = Quote.query.filter(
        Quote.status == "Pending",
        Quote.created_at < cutoff
    ).all()

    count = 0
    for q in expired:
        q.status = "Expired"
        push_notification(
            q.supplier.user_id,
            f"Quote expired - Booking {q.booking.ref}",
            f"Your quote of R{q.amount:,.2f} for {q.booking.route} has expired.",
            type="warning"
        )
        count += 1

    db.session.commit()
    log.info(f"[PIPELINE] Quote expiry - {count} quotes expired")
    return {"expired": count}


# ── Pipeline 5: Invoice Auto-Generation ──────────────────────────────────────

def generate_invoices_for_delivered():
    """
    Extract  - Delivered bookings without an invoice
    Transform - calculate VAT, due date
    Load      - create Invoice + PurchaseOrder records
    """
    from app.models import db, Booking, Invoice, PurchaseOrder
    from datetime import date, timedelta

    delivered = Booking.query.filter_by(status="Delivered").all()
    created   = 0

    for booking in delivered:
        if booking.invoice or not booking.quoted_value:
            continue

        vat_rate   = 0.15
        vat_amount = round(booking.quoted_value * vat_rate, 2)
        total      = round(booking.quoted_value + vat_amount, 2)
        due_date   = date.today() + timedelta(days=30)

        inv = Invoice(
            booking_id     = booking.id,
            invoice_number = f"INV-{datetime.now().year}-{booking.id:05d}",
            amount         = booking.quoted_value,
            vat_amount     = vat_amount,
            total_amount   = total,
            status         = "Unpaid",
            due_date       = due_date,
        )
        db.session.add(inv)

        po = PurchaseOrder(
            booking_id   = booking.id,
            po_number    = f"PO-{datetime.now().year}-{booking.id:05d}",
            gross_amount = booking.quoted_value,
            platform_fee = booking.platform_fee,
            net_payable  = booking.supplier_payout,
            status       = "Pending",
        )
        db.session.add(po)

        from app.services.notifications import push_notification
        push_notification(
            booking.shipper.user_id,
            f"Invoice generated - {inv.invoice_number}",
            f"Invoice for R{total:,.2f} due {due_date}.",
            type="info", ref_type="booking", ref_id=booking.ref
        )
        created += 1

    db.session.commit()
    log.info(f"[PIPELINE] Invoices generated - {created}")
    return {"created": created}


# ── Pipeline 6: Daily Platform Summary ───────────────────────────────────────

def daily_platform_summary():
    """
    Extract  - yesterday's bookings, quotes, revenue
    Transform - aggregate KPIs
    Load      - push summary notification to all admins
    """
    from app.models import db, Booking, User
    from app.services.notifications import push_notification
    from datetime import date, timedelta

    yesterday = date.today() - timedelta(days=1)
    new_bookings = Booking.query.filter(
        db.func.date(Booking.created_at) == yesterday).count()
    delivered    = Booking.query.filter(
        Booking.status == "Delivered",
        db.func.date(Booking.delivered_at) == yesterday).count()
    revenue = db.session.query(db.func.sum(Booking.platform_fee)).filter(
        db.func.date(Booking.delivered_at) == yesterday).scalar() or 0

    admins = User.query.filter_by(role="admin", is_active=True).all()
    for admin in admins:
        push_notification(
            admin.id,
            f"Chart Daily Summary - {yesterday}",
            f"New bookings: {new_bookings} | Delivered: {delivered} | Platform fee earned: R{revenue:,.2f}",
            type="info"
        )

    db.session.commit()
    log.info(f"[PIPELINE] Daily summary sent to {len(admins)} admins")
    return {"new_bookings": new_bookings, "delivered": delivered, "revenue": revenue}


# ── Pipeline 7: Excel Export ──────────────────────────────────────────────────

def generate_excel_report(report_type: str = "bookings"):
    """
    Extract  - full bookings or supplier data from DB
    Transform - structure into tabular format
    Load      - write openpyxl workbook, return bytes
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from app.models import Booking, SupplierProfile
    import io

    wb = openpyxl.Workbook()
    ws = wb.active

    HEADER_FILL = PatternFill("solid", fgColor="155FA0")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)

    def write_headers(ws, headers):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = HEADER_FILL and HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

    if report_type == "bookings":
        ws.title = "Bookings"
        write_headers(ws, ["Ref","Route","Shipper","Supplier","Status",
                            "Value (R)","Platform Fee (R)","Supplier Payout (R)",
                            "Collection Date","Delivered At"])
        for b in Booking.query.order_by(Booking.created_at.desc()).all():
            ws.append([
                b.ref, b.route,
                b.shipper.user.full_name if b.shipper else "",
                b.supplier.company_name if b.supplier else "",
                b.status, b.quoted_value or 0, b.platform_fee or 0,
                b.supplier_payout or 0,
                str(b.collection_date or ""),
                str(b.delivered_at)[:10] if b.delivered_at else ""
            ])

    elif report_type == "suppliers":
        ws.title = "Suppliers"
        write_headers(ws, ["Company","Base City","Status","Score","Total Jobs",
                            "On-Time Rate","Cancellation Rate","Acceptance Rate"])
        for s in SupplierProfile.query.order_by(SupplierProfile.score.desc()).all():
            ws.append([
                s.company_name, s.base_city, s.status, s.score,
                s.total_jobs, s.on_time_rate, s.cancellation_rate, s.acceptance_rate
            ])

    elif report_type == "financial":
        ws.title = "Financial Summary"
        write_headers(ws, ["Ref","Route","Gross (R)","Platform Fee (R)",
                            "Supplier Payout (R)","Status","Date"])
        for b in Booking.query.filter_by(status="Delivered")\
                               .order_by(Booking.delivered_at.desc()).all():
            ws.append([
                b.ref, b.route,
                b.quoted_value or 0, b.platform_fee or 0,
                b.supplier_payout or 0, b.status,
                str(b.delivered_at)[:10] if b.delivered_at else ""
            ])

    # Auto-width columns
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
